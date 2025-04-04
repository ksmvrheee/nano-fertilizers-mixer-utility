import asyncio
import os
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from time import sleep

import flet as ft
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook
from peewee import DoesNotExist, fn

from config import UNIFIED_BUTTON_STYLE
from core.mixture_calculation import calculate_best_mixture
from database import FertilizingMixture, PlantCategory, Plant
from utils.utils import truncate_string


class FullReportPage(ft.Control):
    """
    A page for generating and saving a full report on fertilizer compositions and prices.
    This page provides a GUI to select a file path and generate an Excel report. The report
    contains data about plant categories, fertilization episodes, and the calculated
    optimal fertilizer mixture for each episode.
    """
    def __init__(self, page):
        """Initializes the report generation page."""
        super().__init__()

        self.page = page
        self.ready_to_leave = True

        # obtained path for saving the report
        self.report_filepath = None

        # named labels
        self.heading_container = ft.Container(
            ft.Text(value='Создание полного отчёта',
                    theme_style=ft.TextThemeStyle.HEADLINE_MEDIUM, weight=ft.FontWeight.BOLD),
            margin=ft.margin.only(bottom=2)
        )

        self.filepath_label_container = ft.Container(
            ft.Text(theme_style=ft.TextThemeStyle.TITLE_MEDIUM),
            margin=ft.margin.only(bottom=4)
        )

        # file picker object
        self.file_picker = ft.FilePicker(on_result=self.assess_report_file_path)

        # named buttons
        self.pick_path_button = ft.FilledTonalButton('Выбрать путь сохранения файла', icon=ft.Icons.FOLDER_OPEN,
                                                     style=UNIFIED_BUTTON_STYLE, on_click=self.wrap_file_picking)

        self.pick_path_button_container = ft.Container(self.pick_path_button, margin=ft.margin.only(top=4))

        self.generate_report_button = ft.FilledTonalButton(
            'Создать полный отчёт', style=UNIFIED_BUTTON_STYLE, on_click=self.wrap_report_generation
        )

        # critical error modal
        self.error_dialog = ft.AlertDialog(
            title=ft.Text('Ошибка', color=ft.Colors.RED),
            content=None,
            modal=True,
            scrollable=True,
            actions=[ft.ElevatedButton('Понятно', style=UNIFIED_BUTTON_STYLE, on_click=self.close_error_dialog)]
        )

        # pre-defined views
        self.initial_view = (
            self.heading_container,
            self.pick_path_button_container,
            self.file_picker,
            self.error_dialog
        )

        self.loaded_path_view = (
            self.heading_container,
            self.filepath_label_container,
            self.generate_report_button,
            self.error_dialog
        )

        self.loading_view = (
            ft.Container(
                ft.ProgressRing(),
                width=40, height=40,
                margin=ft.margin.only(top=self.page.height / 3),
                alignment=ft.alignment.center
            ),
            self.error_dialog
        )

        # main content container with the initial view's content
        self.content_container = ft.Container(
            ft.Column(self.initial_view),
            margin=ft.margin.symmetric(vertical=5), alignment=ft.alignment.center
        )

    def build(self):
        """Returns the main container with the initial view."""
        return self.content_container

    def wrap_file_picking(self, e):
        """Opens a file picker dialog for selecting the save path."""
        self.file_picker.save_file(
            dialog_title='Выбор имени и расположения файла отчёта',
            file_name='Отчёт о составах и ценах смесей для нанокапсулирования '
                      f'{datetime.now().strftime("%d-%m-%Y %H-%M-%S")}.xlsx',
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=['xlsx']
        )

    def assess_report_file_path(self, e):
        """Validates the selected file path and updates the UI accordingly."""
        filepath = e.path

        if filepath:
            directory = os.path.dirname(filepath)

            if os.access(directory, os.W_OK):
                self.report_filepath = filepath

                self.filepath_label_container.content.value = f'Путь: {truncate_string(filepath, 50)}'

                self.content_container.content.controls = self.loaded_path_view
                self.content_container.update()

            else:
                self.show_error_dialog(text='Невозможно сохранить файл по данному пути. Выберите другой.')

    def wrap_report_generation(self, e=None):
        """Runs the report generation process asynchronously."""
        self.page.run_task(self._adjust_layout_and_await_report_generation)

    async def _adjust_layout_and_await_report_generation(self, e=None):
        """Updates the UI before and after the report generation process, awaits the report generation."""
        self.content_container.content.controls = self.loading_view
        self.content_container.update()

        self.ready_to_leave = False
        report_saving_result = await asyncio.to_thread(self._generate_and_save_report)
        sleep(0.25)  # waiting to avoid flickering

        self.content_container.content.controls = self.initial_view
        self.content_container.update()

        self.ready_to_leave = True

        if report_saving_result['success']:
            snackbar = ft.SnackBar(ft.Text('Отчёт успешно сохранён.'), duration=1500)
            self.page.overlay.append(snackbar)
            snackbar.open = True
            self.page.update()

    def _generate_and_save_report(self):
        """
        Generates and saves the Excel report.

        :return: a dictionary indicating success or failure.
        """
        wb = Workbook()
        ws = wb.active

        black_bold_font = Font(name="Arial", size=10, bold=True, color="000000")
        black_bold_italic_font = Font(name="Arial", size=10, bold=True, italic=True, color="000000")
        black_regular_font = Font(name="Arial", size=10, color="000000")
        black_italic_font = Font(name="Arial", size=10, italic=True, color="000000")
        red_italic_font = Font(name="Arial", size=10, italic=True, color="FF0000")

        current_cell = ws.cell(row=1, column=1, value='Рассчитанные составы и цены смесей минеральных удобрений')
        current_cell.font = black_bold_font

        current_cell = ws.cell(row=2, column=1, value='для нанокапсулирования')
        current_cell.font = black_bold_font

        current_row = 4

        categories = ((PlantCategory
                       .select()
                       .join(Plant, on=(PlantCategory.id == Plant.category))
                       .group_by(PlantCategory))
                      .having(fn.COUNT(Plant.id) > 0)
                      .order_by(PlantCategory.name))

        mixtures = FertilizingMixture.select()

        if not len(categories):
            current_cell = ws.cell(row=current_row, column=1, value='Данные не найдены.')
            current_cell.font = red_italic_font

            try:
                self.lock_window_closing()
                wb.save(self.report_filepath)

                return {'success': True}

            except (FileNotFoundError, InvalidFileException, PermissionError):
                self.show_error_dialog(text='Не удалось сохранить файл по данному пути, '
                                            'попробуйте другой путь или имя файла.')
                return {'success': False}

            except Exception as err:
                self.show_error_dialog(text=f'Не удалось сохранить файл. Информация об ошибке: {err}.')
                return {'success': False}

            finally:
                self.unlock_window_closing()

        for category in categories:
            current_cell = ws.cell(row=current_row, column=1, value=category.name)
            current_cell.font = black_bold_font
            current_row += 2

            for plant in category.plants.order_by(Plant.name):
                current_cell = ws.cell(row=current_row, column=1, value=plant.name)
                current_cell.font = black_bold_font
                current_row += 1

                for index, fertilizing_episode in enumerate(plant.fertilizing_episodes, 1):
                    current_cell = ws.cell(row=current_row, column=1, value=f'Подкормка №{index}')
                    current_cell.font = black_bold_italic_font
                    current_row += 1

                    episode_nitrogen_mass = fertilizing_episode.nitrogen_mass
                    episode_phosphorus_mass = fertilizing_episode.phosphorus_mass
                    episode_potassium_mass = fertilizing_episode.potassium_mass
                    episode_magnesium_sulfate_mass = fertilizing_episode.magnesium_sulfate_mass

                    magnesium_sulfate_presents = episode_magnesium_sulfate_mass > 0
                    magnesium_sulfate_cost_undefined = False

                    total_mass = episode_nitrogen_mass + episode_phosphorus_mass + episode_potassium_mass

                    episode_masses = {'азота': episode_nitrogen_mass,
                                      'фосфора': episode_phosphorus_mass,
                                      'калия': episode_potassium_mass,
                                      'сульфата магния': episode_magnesium_sulfate_mass}

                    for name, mass in episode_masses.items():
                        if mass:
                            current_cell = ws.cell(row=current_row, column=1, value=f'Масса {name}:')
                            current_cell.font = black_regular_font

                            current_cell = ws.cell(row=current_row, column=2, value=f'{mass} г')
                            current_cell.font = black_regular_font
                            current_row += 1

                    current_cell = ws.cell(row=current_row, column=1, value='Стадия жизненного цикла:')
                    current_cell.font = black_regular_font

                    current_cell = ws.cell(row=current_row, column=2,
                                           value=fertilizing_episode.plant_life_stage_description)
                    current_cell.font = black_regular_font
                    current_row += 1

                    current_cell = ws.cell(row=current_row, column=1, value='Кол-во повторений:')
                    current_cell.font = black_regular_font

                    current_cell = ws.cell(row=current_row, column=2, value=fertilizing_episode.total_repetitions)
                    current_cell.font = black_regular_font
                    current_row += 2

                    current_cell = ws.cell(row=current_row, column=1,
                                           value=f'Рассчитанный оптимальный состав подкормки №{index}')
                    current_cell.font = black_bold_italic_font
                    current_row += 1

                    calculation_result = calculate_best_mixture(
                        nitrogen=episode_nitrogen_mass,
                        phosphorus=episode_phosphorus_mass,
                        potassium=episode_potassium_mass,
                        total_mass=total_mass
                    )

                    if not calculation_result['success']:
                        current_cell = ws.cell(row=current_row, column=1, value=calculation_result['error'])
                        current_cell.font = red_italic_font
                        current_row += 2

                    else:
                        mixture_data = calculation_result['mixture']
                        total_cost = Decimal(str(calculation_result['total_cost']))

                        if magnesium_sulfate_presents:
                            try:
                                magnesium_sulfate_cost = ((FertilizingMixture
                                                           .get(FertilizingMixture.name == 'Сульфат магния')
                                                           .price_per_gram * episode_magnesium_sulfate_mass)
                                                          .quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                                mixture_data.append({
                                    'name': 'Сульфат магния',
                                    'mass_in_grams': episode_magnesium_sulfate_mass,
                                    'cost': magnesium_sulfate_cost
                                })

                                total_cost = ((total_cost + magnesium_sulfate_cost)
                                              .quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

                            except (DoesNotExist, AttributeError):
                                mixture_data.append({'name': 'Сульфат магния'})
                                magnesium_sulfate_cost_undefined = True

                        for mix_unit in mixture_data:
                            mix_unit_name = mix_unit['name']

                            if magnesium_sulfate_cost_undefined and mix_unit_name == 'Сульфат магния':
                                current_cell = ws.cell(row=current_row, column=1, value='Сульфат магния:')
                                current_cell.font = black_regular_font

                                current_cell = ws.cell(row=current_row, column=2,
                                                       value=f'{episode_magnesium_sulfate_mass} г')
                                current_cell.font = black_regular_font

                                current_cell = ws.cell(row=current_row, column=3, value='?')
                                current_cell.font = red_italic_font

                            else:
                                current_cell = ws.cell(row=current_row, column=1, value=mix_unit_name)
                                current_cell.font = black_regular_font

                                current_cell = ws.cell(row=current_row, column=2,
                                                       value=f'{mix_unit["mass_in_grams"]} г')
                                current_cell.font = black_regular_font

                                current_cell = ws.cell(row=current_row, column=3, value=f'{mix_unit["cost"]} руб.')
                                current_cell.font = black_regular_font

                            current_row += 1

                        current_row += 1

                        for key, value in calculation_result['actual_composition'].items():
                            calculation_result['actual_composition'][key] = Decimal(str(value))

                        n_amount_overrun = Decimal(
                            calculation_result['actual_composition']['N'] - episode_nitrogen_mass
                        ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                        p_amount_overrun = Decimal(
                            calculation_result['actual_composition']['P'] - episode_phosphorus_mass
                        ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                        k_amount_overrun = Decimal(
                            calculation_result['actual_composition']['K'] - episode_potassium_mass
                        ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                        for alias, mass in {'азота': n_amount_overrun, 'фосфора': p_amount_overrun,
                                            'калия': k_amount_overrun}.items():
                            if mass:
                                current_cell = ws.cell(row=current_row, column=1, value=f'Превышение массы {alias}:')
                                current_cell.font = black_italic_font

                                current_cell = ws.cell(row=current_row, column=2, value=f'{mass} г')
                                current_cell.font = black_regular_font
                                current_row += 1

                        current_row += 1

                        current_cell = ws.cell(row=current_row, column=1, value='Общая цена:')
                        current_cell.font = black_regular_font

                        if magnesium_sulfate_cost_undefined:
                            current_row += 1
                            current_cell = ws.cell(row=current_row, column=1,
                                                   value='Ошибка: не удалось рассчитать цену, '
                                                         'так как цена за грамм сульфата магния неизвестна.')
                            current_cell.font = red_italic_font

                        else:
                            current_cell = ws.cell(row=current_row, column=2, value=f'{total_cost} руб.')
                            current_cell.font = black_regular_font

                        current_row += 2

        current_cell = ws.cell(row=current_row, column=1, value='Цена за грамм удобрительной смеси')
        current_cell.font = black_bold_font
        current_row += 1

        for mixture in mixtures:
            current_cell = ws.cell(row=current_row, column=1, value=mixture.name)
            current_cell.font = black_regular_font
            current_cell = ws.cell(row=current_row, column=2, value=f'{mixture.price_per_gram} руб.')
            current_cell.font = black_regular_font

            current_row += 1

        # hardcoding the columns' widths for readability
        # because their possible limits are known
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20

        try:
            self.lock_window_closing()
            wb.save(self.report_filepath)

            return {'success': True}

        except (FileNotFoundError, InvalidFileException, PermissionError):
            self.show_error_dialog(text='Не удалось сохранить файл по данному пути, '
                                        'попробуйте другой путь или имя файла.')
            return {'success': False}

        except Exception as err:
            self.show_error_dialog(text=f'Не удалось сохранить файл. Информация об ошибке: {err}.')
            return {'success': False}

        finally:
            self.unlock_window_closing()

    def lock_window_closing(self):
        """Prevents closing the window while saving the report."""
        self.page.window.prevent_close = True
        self.page.update()

    def unlock_window_closing(self):
        """Allows closing the window after report saving."""
        self.page.window.prevent_close = False
        self.page.update()

    def show_error_dialog(self, text):
        """Displays an error dialog with the given text."""
        self.error_dialog.content = ft.Text(text, size=16)
        self.error_dialog.open = True
        self.error_dialog.update()

    def close_error_dialog(self, e=None):
        """Closes the error dialog."""
        self.error_dialog.open = False
        self.error_dialog.update()
